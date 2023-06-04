import sqlite3
import excecoes as exc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import date
from dic_reg_sped_icms_ipi import dic_registro_sped as drs
from dic_reg_sped_contribuicoes import dic_registro_contribuicoes as drc
from hashlib import md5
from os.path import dirname

##### FUNÇÕES PARA VALIDAÇÃO E MANIPULAÇÃO DO ARQUIVO SPED ICMS/IPI EM FORMATO .txt e do arquivo SPED-Contribuições em formtado .txt.#####

# Definição da fonte que irá ser utilizada nos cabeçalhos das colunas. Instância de uma Classe.
fonte_titulo = Font(bold=True, color="1B1B1B")

# Definição das cor de fundo das células que serão utilizadas como título das colunas. Instância de uma Classe.
cor_de_fundo_titulo = PatternFill(fill_type="solid", start_color="E5E5E9")

# Função para manipulçao do arquivo SPED ICMS/IPI em formato .txt.
def manipular_e_processar_arq_sped_icms_ipi(caminho_arquivo):
    dict_arquivo_sped = {}
    # Caminho para encontrar o arquivo, utilizado para salvamento posterior do arquivo processado.
    diretorio = dirname(caminho_arquivo)
    with open(caminho_arquivo, "r", errors="replace") as arquivo:
        # Lê a primeira linha do arquivo, implicando nas leituras posteriores que lerão as linhas após essa.
        l1_arquivo = arquivo.readline()[:]

        # Faz a validação do arquivo SPED ICMS/IPI utilizando a primeira linha do arquivo que contém a informação do registro "|0000|", ou seja, o arquivo será processado desde que contenha este indicativo. Adicionalmente, um arquivo SPED Contribuições poderá ser lido, já que também possui este indicativo de registro, no entanto, as colunas geradas ficarão mal formadas ou não serão encontradas para alguns registros.
        if l1_arquivo[0:6] == "|0000|":
            arquivo_valido = True
        else:
            arquivo_valido = False
        if arquivo_valido is False:
            raise exc.NaoEumArquivoSpedValido("O arquivo selecionado não é um arquivo SPED ICMS/IPI válido.")
        
        # Define o nome do arquivo que será salvo em Excel (.xlsx) na pasta de destino com base em alguns campos.
        lista_l1 = l1_arquivo.split(sep="|")
        del lista_l1[0]
        del lista_l1[-1]

        # Como não poderemos acessar a primeira linha novamente durante a iteração (varredura do arquivo), já adicionaremos ela de início ao dicionário que irá segregar os registros.
        if lista_l1[0] not in dict_arquivo_sped.keys():
            dict_arquivo_sped[lista_l1[0]] = []
            dict_arquivo_sped[lista_l1[0]].append(lista_l1)

        nome_arquivo = "reg_sped_icmsipi_" + str(lista_l1[3]) + "_" + str(lista_l1[4]) + "_" + str(lista_l1[6]) + "_" + str(lista_l1[8]) + ".xlsx"

        # Exclui a primeira linha do arquivo e a lista da primeira linha
        del l1_arquivo, lista_l1

        # Percorrendo linhas no arquivo para listar os dados do arquivo SPED
        for linha in arquivo.readlines():
            lista_temp = linha.split(sep="|")
            del lista_temp[0]
            del lista_temp[-1]
            if lista_temp[0] not in dict_arquivo_sped.keys():
                dict_arquivo_sped[lista_temp[0]] = []
                dict_arquivo_sped[lista_temp[0]].append(lista_temp)
            else:
                dict_arquivo_sped[lista_temp[0]].append(lista_temp)

            # Finalizando a lista de arquivos sped
            if lista_temp[0] == "9999":
                break

        # Criando uma representação de um arquivo em Excel. Por padrão, a primeira aba se chama "Sheet", então excluímos ela e criamos um nova.
        book = Workbook()
        del book["Sheet"]
        book.create_sheet("Indíce")

        # Bloco para manipulação da aba de indíce de registros SPED ICMS IPI. Aqui, acessamos diratamente a planilha, adicionamos as coluns de cabeçalho e estilizamos com negrito e preenchimento de fundo.
        sheet = book["Indíce"]
        sheet.append(["Registros", "Descrição dos Registros"])
        for linha in sheet:
            for cel in linha:
                cel.font = fonte_titulo
                cel.fill = cor_de_fundo_titulo

        # No loop abaixo, ocorre o acesso ao dicionário de dados dos registros SPED ICMS IPI, trazendo o número do registro de acordo com os registros que REALEMENTE estão presentes no arquivo SPED ICMS IPI em formato ".txt". Além disso, caso um registro presente no arquivo ".txt" não estaja presente no diconário de dados de registros (módulo dic_reg_sped_icms_ipi.py), a descrição do registro recebe um texto de atenção "ATENÇÃO!!!(...)". 
        chaves_registros = drs.keys()
        for reg in dict_arquivo_sped:
            if reg in chaves_registros:
                sheet.append([reg, drs[reg]["nome_reg"]])
            else:
                sheet.append([reg, "ATENÇÃO!!! O registro não está incluso no dicionário de dados."])

        # Aplicação de filtro na planilha de indíce.
        sheet.auto_filter.ref = sheet.dimensions

        # Bloco de paginação dos registros SPED ICMS IPI em abas do Excel. O processo é o mesmo do descrito acima, com a diferenciação de que os títulos de colunas são originados diretamente do módulo módulo dic_reg_sped_icms_ipi.py, trazendo os nomes de colunas de acordo com o que existe lá armazenado. Caso o registro processado ainda não tenha sido adicionado ao dicionário de dados, a linha de título de colunas será mal formatada e a aba terá o descritivo de "Descnh - ", para indicar que ali é necessário uma atualização dos dados armazenados.
        for reg in dict_arquivo_sped:
            if reg in chaves_registros:
                # Define o nome da aba, constituída do texto "Reg - " (registro) + o registro referente.
                nome_planilha = "Reg - " + reg
                book.create_sheet(nome_planilha)
                # Estilização das células que representam os títulos de cabeçalho.
                sheet = book[nome_planilha]
                sheet.append(drs[reg]["colunas_reg"])
                for linha in sheet:
                    for cel in linha:
                        cel.font = fonte_titulo
                        cel.fill = cor_de_fundo_titulo

                # Adição das linhas do arquivo SPED ICMS IPI anteriormente separado em um dicionário.
                for linha in dict_arquivo_sped[reg]:
                    sheet.append(linha)

                # Aplicação de filtro na planilha de cada registro.
                sheet.auto_filter.ref = sheet.dimensions

            # Para o caso do registro não ser encontrado no dicionário de armazenamento de dados.
            else:
                nome_planilha = "Descnh - " + reg
                book.create_sheet(nome_planilha)
                sheet = book[nome_planilha]
                sheet.append(["Sem Titulo", "Sem Titulo", "Sem titulo"])
                for linha in dict_arquivo_sped[reg]:
                    sheet.append(linha)

        # Limpa os dados presentes na variável
        del dict_arquivo_sped

        # Define o caminho mais o nome do arqvuivo para salvamento.
        pasta_salvar = diretorio + "/" + nome_arquivo

        # Limpa as variáveis
        del diretorio, nome_arquivo

        # Salva o arquivo no local indicado.
        book.save(pasta_salvar)
        
        # Limpa a variável
        del pasta_salvar

# Função para manipulçao do arquivo SPED-Contribuicões em formato .txt.
def manipular_e_processar_arq_sped_contribuicoes(caminho_arquivo):
    dict_arquivo_sped = {}
    # Caminho para encontrar o arquivo, utilizado para salvamento posterior do arquivo processado.
    diretorio = dirname(caminho_arquivo)
    with open(caminho_arquivo, "r", errors="replace") as arquivo:
        # Lê a primeira linha do arquivo, implicando nas leituras posteriores que lerão as linhas após essa.
        l1_arquivo = arquivo.readline()[:]

        # Faz a validação do arquivo SPED ICMS/IPI utilizando a primeira linha do arquivo que contém a informação do registro "|0000|", ou seja, o arquivo será processado desde que contenha este indicativo. Adicionalmente, um arquivo SPED Contribuições poderá ser lido, já que também possui este indicativo de registro, no entanto, as colunas geradas ficarão mal formadas ou não serão encontradas para alguns registros.
        if l1_arquivo[0:6] == "|0000|":
            arquivo_valido = True
        else:
            arquivo_valido = False
        if arquivo_valido is False:
            raise exc.NaoEumArquivoSpedValido("O arquivo selecionado não é um arquivo SPED válido.")
        
        # Define o nome do arquivo que será salvo em Excel (.xlsx) na pasta de destino com base em alguns campos.
        lista_l1 = l1_arquivo.split(sep="|")
        del lista_l1[0]
        del lista_l1[-1]

        # Como não poderemos acessar a primeira linha novamente durante a iteração (varredura do arquivo), já adicionaremos ela de início ao dicionário que irá segregar os registros.
        if lista_l1[0] not in dict_arquivo_sped.keys():
            dict_arquivo_sped[lista_l1[0]] = []
            dict_arquivo_sped[lista_l1[0]].append(lista_l1)

        print(lista_l1)

        nome_arquivo = "reg_sped_contrib_" + str(lista_l1[5]) + "_" + str(lista_l1[6]) + "_" + str(lista_l1[8]) + "_" + str(lista_l1[9]) + ".xlsx"

        # Exclui a primeira linha do arquivo e a lista da primeira linha
        del l1_arquivo, lista_l1
        # Percorrendo linhas no arquivo para listar os dados do arquivo SPED
        for linha in arquivo.readlines():
            lista_temp = linha.split(sep="|")
            del lista_temp[0]
            del lista_temp[-1]
            if lista_temp[0] not in dict_arquivo_sped.keys():
                dict_arquivo_sped[lista_temp[0]] = []
                dict_arquivo_sped[lista_temp[0]].append(lista_temp)
            else:
                dict_arquivo_sped[lista_temp[0]].append(lista_temp)

            # Finalizando a lista de arquivos sped
            if lista_temp[0] == "9999":
                break

        # Criando uma representação de um arquivo em Excel. Por padrão, a primeira aba se chama "Sheet", então excluímos ela e criamos um nova.
        book = Workbook()
        del book["Sheet"]
        book.create_sheet("Indíce")

        # Bloco para manipulação da aba de indíce de registros SPED ICMS IPI. Aqui, acessamos diratamente a planilha, adicionamos as coluns de cabeçalho e estilizamos com negrito e preenchimento de fundo.
        sheet = book["Indíce"]
        sheet.append(["Registros", "Descrição dos Registros"])
        for linha in sheet:
            for cel in linha:
                cel.font = fonte_titulo
                cel.fill = cor_de_fundo_titulo

        # No loop abaixo, ocorre o acesso ao dicionário de dados dos registros SPED ICMS IPI, trazendo o número do registro de acordo com os registros que REALEMENTE estão presentes no arquivo SPED ICMS IPI em formato ".txt". Além disso, caso um registro presente no arquivo ".txt" não estaja presente no diconário de dados de registros (módulo dic_reg_sped_icms_ipi.py), a descrição do registro recebe um texto de atenção "ATENÇÃO!!!(...)". 
        chaves_registros = drc.keys()
        for reg in dict_arquivo_sped:
            if reg in chaves_registros:
                sheet.append([reg, drc[reg]["nome_reg"]])
            else:
                sheet.append([reg, "ATENÇÃO!!! O registro não está incluso no dicionário de dados."])

        # Aplicação de filtro na planilha de indíce.
        sheet.auto_filter.ref = sheet.dimensions

        # Bloco de paginação dos registros SPED ICMS IPI em abas do Excel. O processo é o mesmo do descrito acima, com a diferenciação de que os títulos de colunas são originados diretamente do módulo módulo dic_reg_sped_icms_ipi.py, trazendo os nomes de colunas de acordo com o que existe lá armazenado. Caso o registro processado ainda não tenha sido adicionado ao dicionário de dados, a linha de título de colunas será mal formatada e a aba terá o descritivo de "Descnh - ", para indicar que ali é necessário uma atualização dos dados armazenados.
        for reg in dict_arquivo_sped:
            if reg in chaves_registros:
                # Define o nome da aba, constituída do texto "Reg - " (registro) + o registro referente.
                nome_planilha = "Reg - " + reg
                book.create_sheet(nome_planilha)
                # Estilização das células que representam os títulos de cabeçalho.
                sheet = book[nome_planilha]
                sheet.append(drc[reg]["colunas_reg"])
                for linha in sheet:
                    for cel in linha:
                        cel.font = fonte_titulo
                        cel.fill = cor_de_fundo_titulo

                # Adição das linhas do arquivo SPED ICMS IPI anteriormente separado em um dicionário.
                for linha in dict_arquivo_sped[reg]:
                    sheet.append(linha)

                # Aplicação de filtro na planilha de cada registro.
                sheet.auto_filter.ref = sheet.dimensions

            # Para o caso do registro não ser encontrado no dicionário de armazenamento de dados.
            else:
                nome_planilha = "Descnh - " + reg
                book.create_sheet(nome_planilha)
                sheet = book[nome_planilha]
                sheet.append(["Sem Titulo", "Sem Titulo", "Sem titulo"])
                for linha in dict_arquivo_sped[reg]:
                    sheet.append(linha)

        # Limpa os dados presentes na variável
        del dict_arquivo_sped

        # Define o caminho mais o nome do arqvuivo para salvamento.
        pasta_salvar = diretorio + "/" + nome_arquivo

        # Limpa as variáveis
        del diretorio, nome_arquivo

        # Salva o arquivo no local indicado.
        book.save(pasta_salvar)
        
        # Limpa a variável
        del pasta_salvar

##### FUNÇÕES PARA VALIDAÇÃO DO PROGRAMA #####

# Função para retornar o texto contendo o mês e o ano atual juntamente com a validade do programa, "Sim" para uso válido e "Não" para uso inválido.
# A função retorna um dado do tipo string.
def texto_data_e_validade():
    with sqlite3.connect("banco_validador.db") as con:
        ano, mes = date.today().year, date.today().month 
        cursor = con.cursor()
        cursor.execute(
            """SELECT senha_autenticada
               FROM   meses_validados 
               WHERE  mes_num = ?
               AND    ano_num = ?;""",
            (mes, ano)
        )
        dado_ma = cursor.fetchone()[0]
        cursor.execute(
            """SELECT cod_hash
               FROM   armaz_senha_mes
               WHERE  mes_num = ?
               AND    ano_num = ?;""",
            (mes, ano)
        )
        dado_hash = cursor.fetchone()[0]
        cursor.close()
        if dado_ma == None:
            return f"{mes}/{ano} - Não"
        if md5(bytes(dado_ma, "utf-8")).hexdigest() == dado_hash:
            return f"{mes}/{ano} - Sim"
        if md5(bytes(dado_ma, "utf-8")).hexdigest() != dado_hash:
            return f"{mes}/{ano} - Não"
        
# Função para buscar a data do último uso do programa.
# Retorna um dado do tipo string.
def busca_data_ultimo_uso():
    with sqlite3.connect("banco_validador.db") as con:
        cursor = con.cursor()
        cursor.execute(
            """SELECT dia, mes, ano
               FROM   registro_processo
               WHERE  sequencia = 1;"""
        )
        dado_cursor = cursor.fetchone()
        if dado_cursor is None:
            cursor.close()
            return ""
        else:
            dado_data = str(dado_cursor[0]) + "/" + str(dado_cursor[1]) + "/" + str(dado_cursor[2])
            cursor.close()
            return dado_data

# Função para informar se o uso atual do programa está válido ou inválido, retornando "S" para programa válido e "N" para prorgama inválido. De outra forma, caso o programa esteja sem o código de autenticação no banco de dado no mês de uso, ele não funcionará durante o processamento do arquivo.
# Retorna um dado do tipo string.
def programa_validado():
    with sqlite3.connect("banco_validador.db") as con:
        ano, mes = date.today().year, date.today().month 
        cursor = con.cursor()
        cursor.execute(
            """SELECT senha_autenticada
               FROM   meses_validados 
               WHERE  mes_num = ?
               AND    ano_num = ?""",
            (mes, ano)
        )
        dado_ma = cursor.fetchone()[0]
        cursor.execute(
            """SELECT cod_hash
               FROM   armaz_senha_mes
               WHERE  mes_num = ?
               AND    ano_num = ?;""",
            (mes, ano)
        )
        dado_hash = cursor.fetchone()[0]
        cursor.close()
        if dado_ma == None:
            return False
        if md5(bytes(dado_ma, "utf-8")).hexdigest() == dado_hash:
            return True
        if md5(bytes(dado_ma, "utf-8")).hexdigest() != dado_hash:
            return False

# Função para confrontar a senha informada no programa com a senha em código hash no banco de dados, caso coincidam, a senha é adicionada à tabela meses_autenticados e o programa funcionará normalmente dentro do mês.
# A função nada retorna, apenas adiciona dados ao banco de dados.
def verificar_e_adicionar_senha(informar_senha):
    if str(informar_senha).strip() == "":
        raise exc.SenhaEmBranco("Não é possível validar uma senha em branco.")
    if informar_senha != "":
        senha = str(informar_senha).strip().lower()
    ano, mes = date.today().year, date.today().month
    with sqlite3.connect("banco_validador.db") as con:
        cursor = con.cursor()
        cursor.execute(
            """SELECT cod_hash
               FROM   armaz_senha_mes
               WHERE  mes_num = ?
               AND    ano_num = ?;""",
            (mes, ano)
        )
        dado_hash = cursor.fetchone()[0]
        if dado_hash == md5(bytes(senha, "utf-8")).hexdigest():
            cursor.execute(
                """UPDATE meses_validados
                   SET    senha_autenticada = ?
                   WHERE  mes_num           = ?
                   AND    ano_num           = ?;""",
                (senha, mes, ano)
            )
            con.commit()
            cursor.close()
        else:
            cursor.close()
            raise exc.SenhaDiferenteDoHash("A senha informada não coincide com a senha para o mês atual.")

# Função para adicionar a data do último processamento de arquivo feito no programa.
# A função nada retorna, apenas adiciona dados ao banco de dados.
def adicionar_data_ultimo_uso():
    data, dia, mes, ano = date.today(), date.today().day, date.today().month, date.today().year
    with sqlite3.connect("banco_validador.db") as con:
        cursor = con.cursor()
        cursor.execute(
            """UPDATE registro_processo
               SET    data_completa = ?,
                      dia           = ?,
                      mes           = ?,
                      ano           = ?
               WHERE  sequencia     = 1;""",
            (data, dia, mes, ano)
        )
        con.commit()
        cursor.close()

# Função para verificar se o mês do ano atual trazido pelo calendário está menor do que o mês do ano do último processamento de arquivo. Caso esteja menor, a função retorna False, caso esteja maior ou igual ao mês do ano, retorna True. Ideal para verificar se a data do sistema foi modificada.
# Retorna um valor booleano (Verdadeiro ou Falso).
def confrontar_data_de_uso_com_data_atual():
    mes = date.today().month
    with sqlite3.connect("banco_validador.db") as con:
        cursor = con.cursor()
        cursor.execute(
            """SELECT mes
               FROM   registro_processo;"""
        )
        dado_mes = cursor.fetchone()
        if dado_mes is not None:
            if mes < dado_mes[0]:
                return False
            else:
                return True
        else:
            return True

# =================================================================
